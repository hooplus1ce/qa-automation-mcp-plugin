import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def render_shadcn_excel(session_data: dict, filepath: str):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "测试步骤"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="0F172A")
    meta_font = Font(name=font_family, size=10, italic=True, color="64748B")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=10, color="0F172A")
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    border_side = Side(style='thin', color='E2E8F0')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title
    ws['A1'] = f"业务测试用例: {session_data.get('flow_name', '未命名')}"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 28
    
    # Meta
    ws['A2'] = f"【系统】: {session_data.get('system_under_test', 'N/A')} | 【描述】: {session_data.get('description', 'N/A')}"
    ws['A2'].font = meta_font
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = ["ID", "操作步骤（业务描述）", "动作", "Iframe 路径", "Playwright 推荐定位器", "测试数据", "预期结果"]
    start_row = 4
    ws.row_dimensions[start_row].height = 24
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Rows
    for i, step in enumerate(session_data.get("steps", [])):
        current_row = start_row + 1 + i
        ws.row_dimensions[current_row].height = 22
        
        l_type = step.get("locator_type")
        l_val = step.get("locator_value")
        l_extra = step.get("locator_extra")
        
        if l_type == "role":
            code_show = f"get_by_role('{l_val}', name='{l_extra or ""}')"
        elif l_type == "test_id":
            code_show = f"get_by_test_id('{l_val}')"
        elif l_type == "label":
            code_show = f"get_by_label('{l_val}')"
        elif l_type == "placeholder":
            code_show = f"get_by_placeholder('{l_val}')"
        elif l_type == "text":
            code_show = f"get_by_text('{l_val}')"
        else:
            code_show = f"locator('{l_val}')"
            
        row_values = [
            step.get("step_number"),
            step.get("description"),
            step.get("action", "").upper(),
            " -> ".join(step.get("frame_path", [])) if step.get("frame_path") else "Main",
            code_show,
            step.get("value") or "-",
            step.get("expected_result") or "系统状态正常响应，无报错异常"
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            
            if col_idx in [1, 3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if i % 2 == 1:
                cell.fill = zebra_fill
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)
        
    wb.save(filepath)
